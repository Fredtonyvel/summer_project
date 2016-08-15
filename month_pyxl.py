import openpyxl
import calendar
from datetime import datetime
from openpyxl.cell import get_column_letter
import os

#loading workbook => summer_projxl.xlsx
wb = openpyxl.load_workbook('month_pyxl.xlsx')
type(wb)
ws = wb.active	#Allows you to write into speadsheet

#Writing values to cells (Title of each column)
sheet = wb.get_sheet_by_name('Sheet1')	#Picked Sheet1

def sheet_title():
	sheet_count = sheet.max_column
	
	sheet[get_column_letter(sheet_count) + str(1)] = 'ID_data'
	sheet_count += 1
	
	sheet[get_column_letter(sheet_count) + str(1)] = 'Name'
	sheet_count += 1
	
	sheet[get_column_letter(sheet_count) + str(1)] = 'Date'
	sheet_count += 1
	
	sheet[get_column_letter(sheet_count) + str(1)] = 'Time'
	sheet_count += 1

	print "Max_column is: ", get_column_letter(sheet_count)

	pass

sheet_title()


'''
sheet['A1'] = 'ID_data'
sheet['B1'] = 'Name'
sheet['C1'] = 'Date'
sheet['D1'] = 'Time'

print sheet['A1'].value
print sheet['B1'].value
print sheet['C1'].value
print sheet['D1'].value
'''





now = datetime.now()

for i in range (2, 11):
	id_data = raw_input("Enter ID #: ")
	sheet['A' + str(i)] = id_data
	print i, ' = ', id_data

print "\n"

for i in range (2, 11):
	name_data = raw_input("Enter student's name: ")
	sheet['B' + str(i)] = name_data
	print i, ' = ', name_data

print "\n"

for i in range (2, 11):
	date = now.strftime("%m/%d/%Y")
	sheet['C' + str(i)] = date
	print i, ' = ', date

print "\n"

for l in range (2,11):
	#time = now.strftime("%H:%M:%S")
	time = now.strftime("%H:%M")
	sheet['D' + str(l)] = time
	print l, ' = ', time

print "\nsheet.max_row =", sheet.max_row
print "sheet.max_column =", get_column_letter(sheet.max_column)

print "\n"




sheet_value = sheet[get_column_letter(sheet.max_column) + str(sheet.max_row)].value
'''
#for i in range (1, sheet.max_row):
if (sheet_value == '12:29'):
	print "Match!"

else:
	print "No Match!"
'''
'''
if (sheet_value == '12:41'):
	sheet_count = sheet.max_column
	sheet_count += 1
	print "New sheet.max_column =", get_column_letter(sheet_count)
else:
	print "Didn't work!"
'''


if (sheet_value == '21:12'):
	sheet_count = sheet.max_column
	sheet_count += 3
	sheet_title()

else:
	print "Didn't work!"


'''
if (sheet_value == '16:20'):
	sheet_count = sheet.max_column
	sheet_count += 3
	print "New sheet.max_column =", get_column_letter(sheet_count)

	sheet[get_column_letter(sheet_count) + str(1)] = 'New_Date'
else:
	print "Didn't work!"
'''


#new_sheet_count += 1
#print "New sheet.max_column =", get_column_letter(sheet_count)


'''
for i in range (2, 11):
	date = now.strftime("%m/%d/%Y")
	sheet['A' + str(i)] = date
	print i, ' = ', date
'''




wb.save('month_pyxl.xlsx')	#Saves updates on excel file

#Using Python to enter commands through CMD prompt
os.system('month_pyxl.xlsx')