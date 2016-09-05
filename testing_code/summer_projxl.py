import openpyxl
import os 

from openpyxl.cell import get_column_letter, column_index_from_string

#Converts letters to numbers
print get_column_letter(1) 

print "\n"

wb = openpyxl.load_workbook('summer_projxl.xlsx')
type(wb)

#Reading in info from Excel
sheet = wb.get_sheet_by_name('Sheet1')
print sheet['A3']
print sheet['A3'].value

print "\n"

#Print out all cells within a column
for i in range(1, 5):
	print (i, sheet.cell(row = i, column = 1).value)

print "\n"

#Writing values to Excel spreadsheet
sheet['A3'] = "hdbfskhjdfb"
print sheet['A3']
print sheet['A3'].value
wb.save('summer_projxl.xlsx')

print "\n"

sheet['A4'] = "Waddup Gary"
print sheet['A4']
print sheet['A4'].value
wb.save('summer_projxl.xlsx')

#Using Python to enter commands through CMD prompt
os.system("summer_projxl.xlsx")