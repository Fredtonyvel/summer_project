import openpyxl
import os
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from datetime import datetime

now = datetime.now()
#print now.strftime("%m/%d/%Y %H:%M:%S")

f=open('ID_Data.txt', 'r')
Id=f.read()
ID = Id.split('\n')
print ID

g=open('Name.txt', 'r')
name=g.read()
Name = name.split('\n')
print Name

#loading workbook => summer_projxl.xlsx
wb = openpyxl.load_workbook('summer_projxl.xlsx')
type(wb)

ws = wb.active	#Allows you to write into speadsheet

#Changing column dimensions
a = ws.column_dimensions['A'].width = 20
b = ws.column_dimensions['B'].width = 20
c = ws.column_dimensions['C'].width = 15
d = ws.column_dimensions['D'].width = 15

#Changing Font colors
#a.font = Font(color=colors.RED)
#b.font = Font(color=colors.BLUE)

e = ws['A1']		
f = ws['B1']
g = ws['C1']
h = ws['D1']

#Applying style changes to each object
e.font = Font(bold=True, size=14)
f.font = Font(bold=True, size=14)	
g.font = Font(bold=True, size=14)
h.font = Font(bold=True, size=14)

#Writing values to cells (Title of each column)
sheet = wb.get_sheet_by_name('Sheet1')	#Picked Sheet1

sheet['A1'] = 'ID Data'
sheet['B1'] = 'Student\'s Name'
sheet['C1'] = 'Date'
sheet['D1'] = 'Time'

id_count = 0
name_count = 0

for p in ID:
	id_count += 1
print "\nNumber of items in ID_Data.txt are: ", id_count

for i in range (2, id_count):
	sheet['A' + str(i)] = ID[i]
	print i, ' = ', ID[i]

print "\n"

for p in ID:
	name_count += 1
print "\nNumber of items in Name.txt are: ", name_count

for j in range (2,name_count):
	sheet['B' + str(j)] = Name[j]
	print j, ' = ', Name[j]

print "\n"

for k in range (2,12):
	date = now.strftime("%m/%d/%Y")
	sheet['C' + str(k)] = date
	print k, ' = ', date

print "\n"

for l in range (2,12):
	time = now.strftime("%H:%M:%S")
	sheet['D' + str(l)] = time
	print l, ' = ', time

'''
#For loop to print out values in each cell
for rowOfCellObjects in sheet['A1': 'D11']:
	for cellObj in rowOfCellObjects:
		print(cellObj.coordinate, cellObj.value)
	print('--- END OF ROW ---\n')
'''

wb.save('summer_projxl.xlsx')	#Saves updates on excel file

#Using Python to enter commands through CMD prompt
os.system('summer_projxl.xlsx')